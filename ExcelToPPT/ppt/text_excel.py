#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Author : Dong-Qing 
# Time : 2018/3/18
from openpyxl import load_workbook
import datetime

work_book = load_workbook(r'F:\we.xlsx')
print(work_book.sheetnames)
sheet_names = work_book.get_sheet_by_name
sheet = work_book['Sheet1']
print(sheet)
s = ""
data = ""
for time in sheet['C']:
    if time.style_id == 3:
        s = time.value
        # data = datetime.time(s)
        x= 1
    # print(data)
    x =1
